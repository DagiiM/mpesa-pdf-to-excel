"""Tests for Excel generation modules."""

import pytest
from unittest.mock import Mock, patch, MagicMock
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.excel_generator.converter import ExcelConverter
from src.excel_generator.summarizer import MonthlySummarizer
from src.excel_generator import ExcelGenerator
from src.utils.exceptions import ExcelGenerationError


class TestExcelConverter:
    """Test cases for ExcelConverter class."""

    def test_init(self, sample_settings):
        """Test ExcelConverter initialization."""
        converter = ExcelConverter(sample_settings)
        assert converter.settings == sample_settings

    def test_create_workbook(self, sample_settings):
        """Test workbook creation."""
        converter = ExcelConverter(sample_settings)
        workbook = converter.create_workbook()
        
        assert isinstance(workbook, Workbook)
        assert "Summary" in workbook.sheetnames
        assert "Transactions" in workbook.sheetnames

    def test_format_currency_cell(self, sample_settings):
        """Test currency cell formatting."""
        converter = ExcelConverter(sample_settings)
        workbook = converter.create_workbook()
        sheet = workbook["Summary"]
        
        # Create a cell and format it
        cell = sheet.cell(row=1, column=1, value=1234.56)
        converter.format_currency_cell(cell)
        
        assert cell.value == 1234.56
        assert cell.number_format == f'"KES"#,##0.00'

    def test_add_summary_data(self, sample_settings, sample_excel_data):
        """Test adding summary data to workbook."""
        converter = ExcelConverter(sample_settings)
        workbook = converter.create_workbook()
        sheet = workbook["Summary"]
        
        converter.add_summary_data(sheet, sample_excel_data["summary"])
        
        # Check that data was added (verify specific cells)
        assert sheet.cell(row=2, column=2).value == sample_excel_data["summary"]["total_credits"]
        assert sheet.cell(row=3, column=2).value == sample_excel_data["summary"]["total_debits"]
        assert sheet.cell(row=4, column=2).value == sample_excel_data["summary"]["net_amount"]

    def test_add_transaction_data(self, sample_settings, sample_transactions):
        """Test adding transaction data to workbook."""
        converter = ExcelConverter(sample_settings)
        workbook = converter.create_workbook()
        sheet = workbook["Transactions"]
        
        converter.add_transaction_data(sheet, sample_transactions)
        
        # Check header row
        assert sheet.cell(row=1, column=1).value == "Date"
        assert sheet.cell(row=1, column=2).value == "Description"
        assert sheet.cell(row=1, column=3).value == "Amount"
        
        # Check first data row
        assert sheet.cell(row=2, column=1).value == sample_transactions[0]["date"]
        assert sheet.cell(row=2, column=2).value == sample_transactions[0]["description"]
        assert sheet.cell(row=2, column=3).value == sample_transactions[0]["amount"]

    def test_add_monthly_breakdown(self, sample_settings, sample_excel_data):
        """Test adding monthly breakdown to workbook."""
        converter = ExcelConverter(sample_settings)
        workbook = converter.create_workbook()
        sheet = workbook["Summary"]
        
        converter.add_monthly_breakdown(sheet, sample_excel_data["monthly_breakdown"])
        
        # Find the monthly breakdown section (should start after row 10)
        monthly_data_found = False
        for row in range(10, 20):
            if sheet.cell(row=row, column=1).value == "2023-01":
                monthly_data_found = True
                assert sheet.cell(row=row, column=2).value == sample_excel_data["monthly_breakdown"]["2023-01"]["credits"]
                break
        
        assert monthly_data_found, "Monthly breakdown data not found in sheet"

    def test_apply_styling(self, sample_settings):
        """Test applying styling to workbook."""
        converter = ExcelConverter(sample_settings)
        workbook = converter.create_workbook()
        
        converter.apply_styling(workbook)
        
        # Check that header cells are styled
        summary_sheet = workbook["Summary"]
        header_cell = summary_sheet.cell(row=1, column=1)
        
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb == "FF4472C4"  # Blue background

    def test_auto_size_columns(self, sample_settings, sample_transactions):
        """Test auto-sizing columns."""
        converter = ExcelConverter(sample_settings)
        workbook = converter.create_workbook()
        sheet = workbook["Transactions"]
        
        converter.add_transaction_data(sheet, sample_transactions)
        converter.auto_size_columns(sheet)
        
        # Check that column dimensions are set (width > 0)
        for column in sheet.columns:
            assert sheet.column_dimensions[column[0].column_letter].width > 0

    def test_save_workbook(self, sample_settings, sample_excel_data, temp_dir):
        """Test saving workbook to file."""
        converter = ExcelConverter(sample_settings)
        output_file = temp_dir / "test_output.xlsx"
        
        converter.save_workbook(sample_excel_data, str(output_file))
        
        assert output_file.exists()
        assert output_file.stat().st_size > 0

    def test_save_workbook_permission_error(self, sample_settings, sample_excel_data):
        """Test saving workbook with permission error."""
        converter = ExcelConverter(sample_settings)
        output_file = "/root/readonly/test_output.xlsx"  # Likely not writable
        
        with pytest.raises(ExcelGenerationError):
            converter.save_workbook(sample_excel_data, output_file)

    def test_format_currency_different_symbol(self, temp_dir):
        """Test currency formatting with different symbol."""
        settings = Mock()
        settings.currency_symbol = "USD"
        converter = ExcelConverter(settings)
        workbook = converter.create_workbook()
        sheet = workbook["Summary"]
        
        cell = sheet.cell(row=1, column=1, value=1234.56)
        converter.format_currency_cell(cell)
        
        assert cell.number_format == f'"USD"#,##0.00'


class TestMonthlySummarizer:
    """Test cases for MonthlySummarizer class."""

    def test_init(self, sample_settings):
        """Test MonthlySummarizer initialization."""
        summarizer = MonthlySummarizer(sample_settings)
        assert summarizer.settings == sample_settings

    def test_calculate_summary(self, sample_settings, sample_transactions):
        """Test summary calculation."""
        summarizer = MonthlySummarizer(sample_settings)
        summary = summarizer.calculate_summary(sample_transactions)
        
        assert summary["total_credits"] == 100.00
        assert summary["total_debits"] == 50.00
        assert summary["net_amount"] == 50.00
        assert summary["transaction_count"] == 2

    def test_calculate_monthly_breakdown(self, sample_settings, sample_transactions):
        """Test monthly breakdown calculation."""
        summarizer = MonthlySummarizer(sample_settings)
        breakdown = summarizer.calculate_monthly_breakdown(sample_transactions)
        
        assert "2023-01" in breakdown
        assert breakdown["2023-01"]["credits"] == 100.00
        assert breakdown["2023-01"]["debits"] == 50.00
        assert breakdown["2023-01"]["net"] == 50.00
        assert breakdown["2023-01"]["transactions"] == 2

    def test_calculate_monthly_breakdown_empty_transactions(self, sample_settings):
        """Test monthly breakdown with empty transactions."""
        summarizer = MonthlySummarizer(sample_settings)
        breakdown = summarizer.calculate_monthly_breakdown([])
        
        assert breakdown == {}

    def test_calculate_monthly_breakdown_multiple_months(self, sample_settings):
        """Test monthly breakdown with transactions across multiple months."""
        summarizer = MonthlySummarizer(sample_settings)
        
        multi_month_transactions = [
            {"date": "2023-01-01", "amount": 100.00, "type": "credit"},
            {"date": "2023-01-15", "amount": -50.00, "type": "debit"},
            {"date": "2023-02-01", "amount": 200.00, "type": "credit"},
            {"date": "2023-02-15", "amount": -75.00, "type": "debit"},
        ]
        
        breakdown = summarizer.calculate_monthly_breakdown(multi_month_transactions)
        
        assert "2023-01" in breakdown
        assert "2023-02" in breakdown
        assert breakdown["2023-01"]["credits"] == 100.00
        assert breakdown["2023-02"]["credits"] == 200.00

    def test_parse_transaction_date(self, sample_settings):
        """Test transaction date parsing."""
        summarizer = MonthlySummarizer(sample_settings)
        
        # Test various date formats
        assert summarizer.parse_transaction_date("01/01/2023") == "2023-01"
        assert summarizer.parse_transaction_date("2023-01-01") == "2023-01"
        assert summarizer.parse_transaction_date("01 Jan 2023") == "2023-01"

    def test_parse_transaction_date_invalid(self, sample_settings):
        """Test parsing invalid transaction dates."""
        summarizer = MonthlySummarizer(sample_settings)
        
        with pytest.raises(ValueError):
            summarizer.parse_transaction_date("invalid_date")

    def test_group_transactions_by_month(self, sample_settings, sample_transactions):
        """Test grouping transactions by month."""
        summarizer = MonthlySummarizer(sample_settings)
        grouped = summarizer.group_transactions_by_month(sample_transactions)
        
        assert "2023-01" in grouped
        assert len(grouped["2023-01"]) == 2

    def test_calculate_month_statistics(self, sample_settings):
        """Test calculating statistics for a month."""
        summarizer = MonthlySummarizer(sample_settings)
        
        month_transactions = [
            {"amount": 100.00, "type": "credit"},
            {"amount": -50.00, "type": "debit"},
            {"amount": 200.00, "type": "credit"},
        ]
        
        stats = summarizer.calculate_month_statistics(month_transactions)
        
        assert stats["credits"] == 300.00
        assert stats["debits"] == 50.00
        assert stats["net"] == 250.00
        assert stats["transactions"] == 3


class TestExcelGenerator:
    """Test cases for ExcelGenerator class."""

    def test_init(self, sample_settings):
        """Test ExcelGenerator initialization."""
        generator = ExcelGenerator(sample_settings)
        assert generator.settings == sample_settings
        assert generator.converter is not None
        assert generator.summarizer is not None

    def test_generate_excel_report(self, sample_settings, sample_transactions, temp_dir):
        """Test complete Excel report generation."""
        generator = ExcelGenerator(sample_settings)
        output_file = temp_dir / "test_report.xlsx"
        
        result = generator.generate_excel_report(sample_transactions, str(output_file))
        
        assert result == str(output_file)
        assert Path(output_file).exists()
        assert Path(output_file).stat().st_size > 0

    def test_generate_excel_report_with_empty_transactions(self, sample_settings, temp_dir):
        """Test Excel report generation with empty transactions."""
        generator = ExcelGenerator(sample_settings)
        output_file = temp_dir / "empty_report.xlsx"
        
        result = generator.generate_excel_report([], str(output_file))
        
        assert result == str(output_file)
        assert Path(output_file).exists()

    def test_generate_excel_report_file_creation_error(self, sample_settings, sample_transactions):
        """Test Excel report generation with file creation error."""
        generator = ExcelGenerator(sample_settings)
        output_file = "/nonexistent/path/report.xlsx"
        
        with pytest.raises(ExcelGenerationError):
            generator.generate_excel_report(sample_transactions, output_file)

    def test_generate_excel_report_with_metadata(self, sample_settings, sample_transactions, temp_dir):
        """Test Excel report generation with metadata."""
        generator = ExcelGenerator(sample_settings)
        output_file = temp_dir / "metadata_report.xlsx"
        
        # Add metadata to transactions
        transactions_with_metadata = {
            "transactions": sample_transactions,
            "metadata": {
                "source_file": "test.pdf",
                "processing_date": "2023-01-01",
                "total_records": len(sample_transactions),
            }
        }
        
        result = generator.generate_excel_report(
            transactions_with_metadata["transactions"], 
            str(output_file),
            metadata=transactions_with_metadata["metadata"]
        )
        
        assert result == str(output_file)
        assert Path(output_file).exists()

    def test_validate_output_path(self, sample_settings, temp_dir):
        """Test output path validation."""
        generator = ExcelGenerator(sample_settings)
        
        # Valid path
        valid_path = temp_dir / "test.xlsx"
        assert generator.validate_output_path(str(valid_path)) is True
        
        # Invalid path (directory doesn't exist)
        invalid_path = "/nonexistent/directory/test.xlsx"
        assert generator.validate_output_path(invalid_path) is False
        
        # Invalid extension
        invalid_ext = temp_dir / "test.txt"
        assert generator.validate_output_path(str(invalid_ext)) is False

    def test_get_default_output_filename(self, sample_settings):
        """Test default output filename generation."""
        generator = ExcelGenerator(sample_settings)
        
        filename = generator.get_default_output_filename()
        
        assert filename.endswith(".xlsx")
        assert "bank_statement" in filename.lower()