"""Excel conversion utilities for bank statement data."""

import os
from datetime import datetime
from typing import Dict, List, Optional, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.config.settings import REPORTS_DIR, EXCEL_OUTPUT_FORMAT, INCLUDE_METADATA, CURRENCY_SYMBOL
from src.pdf_processor.extractor import TransactionData
from src.utils.logger import get_logger
from src.utils.validators import ValidationError, validate_directory_path


class ExcelConversionError(Exception):
    """Custom exception for Excel conversion errors."""
    pass


class ExcelConverter:
    """Handles conversion of transaction data to Excel format."""
    
    def __init__(self) -> None:
        """Initialize Excel converter."""
        self.logger = get_logger(__name__)
        
        # Define Excel styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.currency_format = f'[{CURRENCY_SYMBOL}] #,##0.00'
        self.date_format = 'YYYY-MM-DD'
    
    def generate_filename(
        self,
        base_name: str,
        suffix: Optional[str] = None,
        timestamp: bool = True
    ) -> str:
        """Generate Excel filename with timestamp.
        
        Args:
            base_name: Base filename.
            suffix: Optional suffix to add.
            timestamp: Whether to include timestamp.
            
        Returns:
            Generated filename.
        """
        timestamp_str = ""
        if timestamp:
            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        parts = [base_name]
        if suffix:
            parts.append(suffix)
        if timestamp_str:
            parts.append(timestamp_str)
        
        filename = "_".join(parts)
        return f"{filename}.{EXCEL_OUTPUT_FORMAT}"
    
    def transactions_to_dataframe(self, transactions: List[TransactionData]) -> pd.DataFrame:
        """Convert transactions list to pandas DataFrame.
        
        Args:
            transactions: List of TransactionData objects.
            
        Returns:
            pandas DataFrame with transaction data.
        """
        data = []
        for transaction in transactions:
            row = {
                "Date": transaction.date,
                "Description": transaction.description,
                "Debit": float(transaction.debit) if transaction.debit else 0.0,
                "Credit": float(transaction.credit) if transaction.credit else 0.0,
                "Balance": float(transaction.balance) if transaction.balance else None,
                "Reference": transaction.reference or "",
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Convert date column to datetime
        if not df.empty and 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        
        # Sort by date
        if not df.empty and 'Date' in df.columns:
            df = df.sort_values('Date')
        
        return df
    
    def create_transactions_sheet(
        self,
        workbook: Workbook,
        transactions_df: pd.DataFrame,
        sheet_name: str = "Transactions"
    ) -> None:
        """Create transactions sheet in workbook.
        
        Args:
            workbook: Excel workbook object.
            transactions_df: DataFrame with transaction data.
            sheet_name: Name for the sheet.
        """
        if transactions_df.empty:
            self.logger.warning("No transaction data to write to Excel")
            return
        
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Write headers
        headers = list(transactions_df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Write data
        for row_num, row in enumerate(dataframe_to_rows(transactions_df, index=False, header=False), 2):
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                
                # Apply formatting based on column type
                if col_num == 1 and isinstance(value, datetime):  # Date column
                    cell.number_format = self.date_format
                elif col_num in [3, 4, 5] and isinstance(value, (int, float)):  # Amount columns
                    cell.number_format = self.currency_format
        
        # Auto-adjust column widths using column letters directly
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 0
            column_letter = chr(64 + col_idx)  # Convert to A, B, C, etc.
            
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Created transactions sheet with {len(transactions_df)} rows")
    
    def create_metadata_sheet(
        self,
        workbook: Workbook,
        metadata: Dict[str, Any],
        sheet_name: str = "Metadata"
    ) -> None:
        """Create metadata sheet in workbook.
        
        Args:
            workbook: Excel workbook object.
            metadata: Dictionary with metadata information.
            sheet_name: Name for the sheet.
        """
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Write metadata
        row_num = 1
        for key, value in metadata.items():
            worksheet.cell(row=row_num, column=1, value=str(key))
            worksheet.cell(row=row_num, column=2, value=str(value))
            
            # Format header row
            if row_num == 1:
                cell = worksheet.cell(row=row_num, column=1)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                
                cell = worksheet.cell(row=row_num, column=2)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            
            row_num += 1
        
        # Auto-adjust column widths using column letters directly
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 0
            column_letter = chr(64 + col_idx)  # Convert to A, B, C, etc.
            
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Created metadata sheet with {len(metadata)} items")
    
    def convert_to_excel(
        self,
        transactions: List[TransactionData],
        output_path: Optional[str] = None,
        filename: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """Convert transactions to Excel file.
        
        Args:
            transactions: List of TransactionData objects.
            output_path: Optional output directory path.
            filename: Optional filename for output file.
            metadata: Optional metadata to include.
            
        Returns:
            Path to created Excel file.
            
        Raises:
            ExcelConversionError: If conversion fails.
        """
        try:
            # Validate output directory
            if output_path is None:
                output_path = REPORTS_DIR
            
            validate_directory_path(output_path)
            
            # Generate filename if not provided
            if filename is None:
                filename = self.generate_filename("bank_statement")
            
            # Ensure filename has correct extension
            if not filename.endswith(f".{EXCEL_OUTPUT_FORMAT}"):
                filename = f"{filename}.{EXCEL_OUTPUT_FORMAT}"
            
            full_path = os.path.join(output_path, filename)
            
            # Convert transactions to DataFrame
            transactions_df = self.transactions_to_dataframe(transactions)
            
            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            
            # Create transactions sheet
            self.create_transactions_sheet(workbook, transactions_df)
            
            # Create metadata sheet if enabled
            if INCLUDE_METADATA and metadata:
                self.create_metadata_sheet(workbook, metadata)
            
            # Save workbook
            workbook.save(full_path)
            workbook.close()
            
            self.logger.info(f"Excel file created successfully: {full_path}")
            return full_path
            
        except ValidationError as e:
            raise ExcelConversionError(f"Validation error: {str(e)}")
        except Exception as e:
            raise ExcelConversionError(f"Failed to convert to Excel: {str(e)}")
    
    def create_summary_excel(
        self,
        summary_data: Dict[str, Any],
        transactions: List[TransactionData],
        output_path: Optional[str] = None,
        filename: Optional[str] = None
    ) -> str:
        """Create Excel file with summary and transaction data.
        
        Args:
            summary_data: Dictionary with summary information.
            transactions: List of TransactionData objects.
            output_path: Optional output directory path.
            filename: Optional filename for output file.
            
        Returns:
            Path to created Excel file.
            
        Raises:
            ExcelConversionError: If creation fails.
        """
        try:
            # Validate output directory
            if output_path is None:
                output_path = REPORTS_DIR
            
            validate_directory_path(output_path)
            
            # Generate filename if not provided
            if filename is None:
                filename = self.generate_filename("bank_summary")
            
            # Ensure filename has correct extension
            if not filename.endswith(f".{EXCEL_OUTPUT_FORMAT}"):
                filename = f"{filename}.{EXCEL_OUTPUT_FORMAT}"
            
            full_path = os.path.join(output_path, filename)
            
            # Convert transactions to DataFrame
            transactions_df = self.transactions_to_dataframe(transactions)
            
            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            
            # Create summary sheet
            self.create_summary_sheet(workbook, summary_data)
            
            # Create transactions sheet
            self.create_transactions_sheet(workbook, transactions_df)
            
            # Save workbook
            workbook.save(full_path)
            workbook.close()
            
            self.logger.info(f"Summary Excel file created successfully: {full_path}")
            return full_path
            
        except ValidationError as e:
            raise ExcelConversionError(f"Validation error: {str(e)}")
        except Exception as e:
            raise ExcelConversionError(f"Failed to create summary Excel: {str(e)}")
    
    def create_summary_sheet(
        self,
        workbook: Workbook,
        summary_data: Dict[str, Any],
        sheet_name: str = "Summary"
    ) -> None:
        """Create summary sheet in workbook.
        
        Args:
            workbook: Excel workbook object.
            summary_data: Dictionary with summary information.
            sheet_name: Name for the sheet.
        """
        worksheet = workbook.create_sheet(title=sheet_name)
        row_num = 1
        
        # Title
        title_cell = worksheet.cell(row=row_num, column=1, value="Bank Statement Summary")
        title_cell.font = Font(bold=True, size=16)
        worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        row_num += 2
        
        # Overall Summary Section
        self._write_overall_summary(worksheet, summary_data, row_num)
        row_num += 8
        
        # Analysis Period Section
        if "analysis_period" in summary_data:
            self._write_analysis_period(worksheet, summary_data["analysis_period"], row_num)
            row_num += 5
        
        # Monthly Summaries Section
        if "monthly_summaries" in summary_data:
            row_num = self._write_monthly_summaries(worksheet, summary_data["monthly_summaries"], row_num)
        
        # Auto-adjust column widths
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 0
            column_letter = chr(64 + col_idx)  # Convert to A, B, C, etc.
            
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Created summary sheet with comprehensive data")
    
    def _write_overall_summary(
        self,
        worksheet,
        summary_data: Dict[str, Any],
        start_row: int
    ) -> None:
        """Write overall summary section.
        
        Args:
            worksheet: Excel worksheet object.
            summary_data: Summary data dictionary.
            start_row: Starting row number.
        """
        # Section header
        header_cell = worksheet.cell(row=start_row, column=1, value="Overall Summary")
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = self.header_alignment
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        
        # Overall totals
        if "overall_totals" in summary_data:
            totals = summary_data["overall_totals"]
            row = start_row + 2
            
            worksheet.cell(row=row, column=1, value="Total Transactions:")
            worksheet.cell(row=row, column=2, value=summary_data.get("total_transactions", 0))
            row += 1
            
            worksheet.cell(row=row, column=1, value="Total Credits:")
            worksheet.cell(row=row, column=2, value=f"{CURRENCY_SYMBOL} {totals.get('total_credits', 0):,.2f}")
            row += 1
            
            worksheet.cell(row=row, column=1, value="Total Debits:")
            worksheet.cell(row=row, column=2, value=f"{CURRENCY_SYMBOL} {totals.get('total_debits', 0):,.2f}")
            row += 1
            
            worksheet.cell(row=row, column=1, value="Net Amount:")
            worksheet.cell(row=row, column=2, value=f"{CURRENCY_SYMBOL} {totals.get('net_amount', 0):,.2f}")
            row += 1
            
            worksheet.cell(row=row, column=1, value="Average Monthly Transactions:")
            worksheet.cell(row=row, column=2, value=f"{summary_data.get('average_monthly_transactions', 0):.1f}")
    
    def _write_analysis_period(
        self,
        worksheet,
        analysis_period: Dict[str, Any],
        start_row: int
    ) -> None:
        """Write analysis period section.
        
        Args:
            worksheet: Excel worksheet object.
            analysis_period: Analysis period data.
            start_row: Starting row number.
        """
        # Section header
        header_cell = worksheet.cell(row=start_row, column=1, value="Analysis Period")
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = self.header_alignment
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        
        # Period details
        row = start_row + 2
        worksheet.cell(row=row, column=1, value="Start Date:")
        worksheet.cell(row=row, column=2, value=analysis_period.get("start_date", "N/A"))
        row += 1
        
        worksheet.cell(row=row, column=1, value="End Date:")
        worksheet.cell(row=row, column=2, value=analysis_period.get("end_date", "N/A"))
        row += 1
        
        worksheet.cell(row=row, column=1, value="Total Days:")
        worksheet.cell(row=row, column=2, value=analysis_period.get("total_days", 0))
    
    def _write_monthly_summaries(
        self,
        worksheet,
        monthly_summaries: Dict[str, Dict[str, Any]],
        start_row: int
    ) -> int:
        """Write monthly summaries section.
        
        Args:
            worksheet: Excel worksheet object.
            monthly_summaries: Monthly summary data.
            start_row: Starting row number.
            
        Returns:
            Next available row number.
        """
        # Section header
        header_cell = worksheet.cell(row=start_row, column=1, value="Monthly Breakdown")
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = self.header_alignment
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        
        # Table headers
        row = start_row + 2
        headers = ["Month", "Transactions", "Total Credits", "Total Debits", "Net Amount"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Monthly data
        row += 1
        for month, data in sorted(monthly_summaries.items()):
            worksheet.cell(row=row, column=1, value=data.get("month", month))
            worksheet.cell(row=row, column=2, value=data.get("transaction_count", 0))
            worksheet.cell(row=row, column=3, value=f"{CURRENCY_SYMBOL} {data.get('total_credits', 0):,.2f}")
            worksheet.cell(row=row, column=4, value=f"{CURRENCY_SYMBOL} {data.get('total_debits', 0):,.2f}")
            worksheet.cell(row=row, column=5, value=f"{CURRENCY_SYMBOL} {data.get('net_amount', 0):,.2f}")
            row += 1
        
        return row + 2